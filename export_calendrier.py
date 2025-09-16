#!/usr/bin/env python

"""
@File    :   export_calendrier.py
@Time    :   2023/09/23 11:38:01
@Version :   1.0
@Desc    :   recuperation du calendrier sur cyu.fr
"""

# Importer les modules nécessaires pour la locale, les opérations sur fichiers, la gestion des dates, la journalisation, les délais temporels, la sérialisation des données et l'exécution de sous-processus
import locale
import os
import datetime
import logging
import time
import pickle
import subprocess

# Importer le client proxy KeePassXC pour la récupération sécurisée des identifiants
import keepassxc_proxy_client
import keepassxc_proxy_client.protocol

# Importer Selenium pour l'automatisation web avec Firefox
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options

# Importer OpenPyXL pour la création et le style des fichiers Excel
from openpyxl import Workbook
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
from openpyxl.styles import Font

# Importer Tkinter pour les dialogues d'interaction utilisateur
import tkinter.messagebox as msg

# Nom d'étudiant par défaut
etudiant="Gabriel"


# Classe de base pour les identifiants vides
class Dict2ClassEmpty(object):
    def __init__(self):
        self.login = ""
        self.password = ""


# Classe pour convertir un dictionnaire en objet avec des attributs
class Dict2Class(Dict2ClassEmpty):
    def __init__(self, my_dict):
        super().__init__()
        for key in my_dict:
            setattr(self, key, my_dict[key])


# Fonction pour récupérer les identifiants depuis KeePassXC pour une URL donnée
def get_credential(url):
    """
    Récupère les identifiants depuis KeepassXC
    :param url: L'URL pour laquelle récupérer les identifiants
    :return: Objet identifiant ou objet vide si non trouvé
    """
    global etudiant
    # Établir la connexion au proxy KeePassXC
    connection = keepassxc_proxy_client.protocol.Connection()
    connection.connect()
    _home = os.environ.get("HOME", "")
    # Vérifier si le fichier d'association existe, le créer sinon
    if not os.path.exists(os.path.join(_home, ".ssh", "python_keepassxc")):
        connection.associate()
        name, public_key = connection.dump_associate()
        print("Got connection named '", name, "' with key", public_key)
        with open(os.path.join(_home, ".ssh", "python_keepassxc"), "wb") as fic:
            fic.write(public_key)
    # Charger l'association existante
    with open(os.path.join(_home, ".ssh", "python_keepassxc"), "rb") as fic:
        public_key = fic.read()
    connection.load_associate("python", public_key)
    connection.test_associate()
    # Récupérer les identifiants pour l'URL
    credentials = connection.get_logins(url)
    if credentials:
        for gugusse in credentials:
            wanted_etudiant = "e-%s" % etudiant[0].lower()
            if wanted_etudiant in gugusse["login"]:
                credential = Dict2Class(gugusse)
                logging.debug(credential.login)
                return credential
    return Dict2ClassEmpty()


# Bloc d'exécution principal
if __name__ == "__main__":
    # Demander à l'utilisateur de sélectionner l'étudiant
    for choix in ["Gabriel", "Louis-Joseph"]:
        if msg.askyesno("Étudiant?", choix):
            etudiant = choix
            break
    # Configurer la journalisation
    logging.basicConfig(encoding="utf-8", level=logging.INFO)
    my_creds = Dict2ClassEmpty()
    # Définir la locale en français
    locale.setlocale(locale.LC_ALL, "fr_FR.utf8")
    # Configurer les options de Firefox
    options = Options()
    options.binary_location = "/usr/lib/firefox/firefox"
    driver = None

    # Générer la liste des jours de la semaine en français
    jours = list(datetime.date(2001, 1, i).strftime("%A") for i in range(1, 7))
    # Calculer la date de début (lundi prochain)
    start_date = datetime.date.today()
    if start_date.weekday() != 0:
        start_date += datetime.timedelta(days=7 - start_date.weekday())
    dates = jours
    # Ajouter les dates aux jours de la semaine
    for ind_i, _ in enumerate(jours):
        n_date = start_date + datetime.timedelta(days=ind_i)
        dates[ind_i] += n_date.strftime(" %d/%m")
    logging.debug(dates)
    # Vérifier si les données mises en cache existent et sont récentes (dans les 6 jours)
    if os.path.exists("e_d_t_%s.dat" % etudiant.lower()):
        # vérification de la date d'export
        now = datetime.datetime.now()
        debut = now - datetime.timedelta(days=6)
        st = os.stat("e_d_t_%s.dat" % etudiant.lower())
        date_f = datetime.datetime.fromtimestamp(st.st_mtime)
        if date_f < debut:
            os.remove("e_d_t_%s.dat" % etudiant.lower())
    # Charger les données mises en cache si disponibles
    e_d_t = None
    if os.path.exists("e_d_t_%s.dat" % etudiant.lower()):
        # export encore valide
        with open("e_d_t_%s.dat" % etudiant.lower(), "rb") as driver_dat:
            e_d_t = pickle.load(driver_dat)
    else:
        # Aucune donnée mise en cache disponible, récupérer depuis cyu.fr
        my_creds = get_credential("https://cyu.fr/")
        my_creds2 = get_credential("https://services-web.cyu.fr/calendar/cal")
        
        # Initialiser le pilote Firefox
        driver = webdriver.Firefox()
        URL = "https://services-web.cyu.fr/calendar/"
        LOGIN_ROUTE = "LdapLogin/"
        DATA_ROUTE = "cal?vt=agendaWeek&dt=%s&et=student&fid0=%s" % (start_date.isoformat(), my_creds2.password)

        # Naviguer vers la page de connexion
        driver.get(URL + LOGIN_ROUTE)

        # Saisir les identifiants et se connecter
        u_name = driver.find_element(By.ID, "Name")
        u_name.send_keys(my_creds.login)
        u_password = driver.find_element(By.ID, "Password")
        u_password.send_keys(my_creds.password)
        driver.find_element(By.CSS_SELECTOR, ".loginBtn").click()
        # Naviguer vers la page des données du calendrier
        driver.get(URL + DATA_ROUTE)
        time.sleep(1)

        # Initialiser la structure de données pour l'emploi du temps
        if e_d_t is None:
            e_d_t = [["Horaire", "Cours / TD", "Salle", "Prof.", "          "]]
        # Analyser les données du calendrier depuis la page web
        for ind_i, col in enumerate(driver.find_elements(By.CSS_SELECTOR, ".fc-content-col")):
            e_d_t.append([dates[ind_i], "", "", ""])
            for output in col.find_elements(By.XPATH, value="(.//*[contains(@class, 'fc-content')])"):
                if output.text:
                    # Extraire l'heure et le contenu
                    heure_i, sans_lignes = output.text.split("\n", 1)
                    prof_i = ""
                    try:
                        # Analyser pour le site CH
                        contenu_i, suite_i = sans_lignes.split("\nCH", 1)
                        salle_i = "CH" + suite_i.split("p\n")[0] + "p"
                        salle_i = "\n".join(salle_i.split(" ", 1))
                        salle_i = "\n".join(salle_i.rsplit(" ", 1))
                    except ValueError:
                        try:
                            # Gérer l'analyse alternative
                            sans_lignes_1 = sans_lignes.split("\n", 1)[0]
                            sans_lignes_2 = sans_lignes.split(")\n", 1)[1]
                            sans_lignes = "\n".join([sans_lignes_1, sans_lignes_2])
                        except IndexError:
                            pass
                        contenu_i = ""
                        salle_i = ""
                        suite_i = ""
                        # Essayer d'analyser pour les sites NEU ou TUR
                        for site in ["NEU", "TUR"]:
                            try:
                                contenu_i, suite_i = sans_lignes.split("\n%s" % site, 1)
                                salle_i = site + suite_i.split("p\n")[0] + "p"
                                break
                            except ValueError:
                                continue
                        if not contenu_i:
                            if "à distance" in sans_lignes:
                                # Gérer les cours à distance
                                reconstruction = sans_lignes.split("\n")
                                reconstruction.pop()
                                prof_i = reconstruction.pop()
                                contenu_i = "\n".join(reconstruction)
                                salle_i = "à la maison"
                            else:
                                # Analyse par défaut
                                contenu_i, suite_i = sans_lignes.split("\n", 1)
                                salle_i = suite_i.split("p\n")[0] + "p"
                    try:
                        # Extraire le nom du professeur
                        if salle_i != "à la maison":
                            prof_i = suite_i.split("p\n")[1].split('\n')[0]
                    except IndexError:
                        pass
                    # Ajouter les données analysées à l'emploi du temps
                    e_d_t.append(["- %s" % heure_i, contenu_i, salle_i, prof_i, ""])
        # Mettre en cache les données analysées
        with open("e_d_t_%s.dat" % etudiant.lower(), "wb") as driver_dat:
            # Suppression du faux positif de PyTypeChecker dans PyCharm
            # noinspection PyTypeChecker
            pickle.dump(e_d_t, driver_dat)

    # Créer un classeur et une feuille Excel
    wb = Workbook()
    ws = wb.active
    blueFill = PatternFill(start_color="FFAAAAFF", end_color="FFAAAAFF", fill_type="solid")
    if ws is None:
        exit(1)
    # Remplir la feuille avec les données d'emploi du temps
    for ligne in e_d_t:
        ws.append(ligne)
    # Mettre en gras la ligne d'en-tête
    for row in ws.rows:
        for cell in row:
            cell.font = Font(bold=True)
        break

    # Configurer la mise en page
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
    dims = {}
    # Calculer les largeurs de colonnes et appliquer le style
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max([len(ligne) for ligne in str(cell.value).splitlines()]) + 5
            if not row[2].value:
                cell.fill = blueFill
                cell.font = Font(bold=True)
    # Définir les largeurs de colonnes
    for col, value in dims.items():
        ws.column_dimensions[chr(64 + col)].width = value
    # Définir les paramètres d'impression selon l'étudiant
    if etudiant == "Gabriel":
        Worksheet.set_printer_settings(ws, paper_size=9, orientation="landscape")
    else:
        Worksheet.set_printer_settings(ws, paper_size=9, orientation="portrait")
    # Sauvegarder le fichier Excel
    wb.save("e_d_t_%s.xlsx" % etudiant.lower())
    # Ouvrir le fichier Excel
    subprocess.run(["open", "e_d_t_%s.xlsx" % etudiant.lower()], check=False)
    # Fermer le navigateur s'il a été ouvert
    if driver:
        driver.close()
